import pymysql
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import warnings
import json

# admin_file = open('/home/ssnoc/core_negrp/ApplicationInfo.json')
admin_file = open('/Users/yuyungwang/Documents/GitHub/core_negrp/ApplicationInfo.json')
admin_dict = json.load(admin_file)

# sql_dict = admin_dict['mysql_db']
sql_dict = admin_dict['local_db']

sql_host = sql_dict['host']
sql_user = sql_dict['user']
sql_passwd = sql_dict['passwd']
sql_charset = sql_dict['charset']

LIST_MAX_COUNT = 20
DETAIL_MAX_LEN = 700

# mylogger = Logger.make_logger('local_db')

class DBHandler:
    def __init__(self):
        pass

    def connectiondb(dbname):
        try:
            con = pymysql.connect(host = sql_host,
                       user = sql_user,
                       password = sql_passwd,
                       db = dbname,
                       charset = 'utf8')
            return con
        except Exception as e:
            print(str(e))

    def insert_ipmap_from_excel(self,excel_filename):
        try:
            with DBHandler.connectiondb('TraceRoute') as conn:
                with conn.cursor() as cur:
                    sql = 'INSERT INTO IPMAP (IP,Equip,date) VALUES(%s, %s, CURDATE());'
                    # sql = "INSERT INTO IPMAP (IP,Equip,date)  VALUES(%s, %s, CURDATE()) WHERE NOT EXISTS (SELECT * FROM IPMAP WHERE IP = %s AND Equip = %s);" \
                    # "UPDATE IPMAP SET Equip=%s WHERE IP = %s AND Equip IS NOT %s;"
                    wb = load_workbook(excel_filename, data_only=True)
                    ws = wb['ip_map']
                    #warnings.simplefilter("default")
        
                    iter_rows = iter(ws.rows)
                    next(iter_rows)
                    
                    for row in iter_rows:
                        tmp1 = row[0].value
                        tmp2 = row[1].value
                        cur.execute(sql,(tmp1,tmp2,tmp1,tmp2,tmp2,tmp1,tmp2))
                    conn.commit()

                    
                    wb.close()
                    conn.close()
        except Exception as e:
            print(str(e))

    # def get_ip_list(input):
    #     try:
    #         result = []
    #         sql = f"SELECT DISTINCT * from ip_host where ip like '%{input}%' OR hostname like '%{input}%' OR equip_name like '%{input}%' limit 50;"
    #         with DBHandler.connectiondb('coreops') as conn:
    #             with conn.cursor() as cur:
    #                 # num_fields = len(cur.description)
    #                 # field_names = [i[0] for i in cur.description]
    #                 # print(field_names)
    #                 cur.execute(sql)
    #                 result = cur.fetchall()
    #         result = [list(result[x]) for x in range(len(result))]
    #         return result
    #     except Exception as e:
    #         print(str(e))
            
if __name__=='__main__':
    mydb = DBHandler()
    # DBHandler.insert_ipmap_from_excel('/home/ssnoc/core_negrp/ip_map.xlsx')
    mydb.insert_ipmap_from_excel('/Users/yuyungwang/Documents/GitHub/core_negrp/ip_map.xlsx')
